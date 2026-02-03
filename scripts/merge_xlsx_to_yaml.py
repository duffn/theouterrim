#!/usr/bin/env python3
"""
Merge XLSX spreadsheet data into YAML files.

Reads XLSX files from google_sheets/, compares against _data/*.yaml,
adds missing items, and produces a diff report for existing items with
differing data.

Usage:
    uv run --with openpyxl,ruamel.yaml scripts/merge_xlsx_to_yaml.py [--dry-run] [--verbose]
"""

import argparse
import datetime
import os
import re
import sys
import uuid
from pathlib import Path

import openpyxl
from ruamel.yaml import YAML

# ---------------------------------------------------------------------------
# Paths
# ---------------------------------------------------------------------------
BASE_DIR = Path(__file__).resolve().parent.parent
DATA_DIR = BASE_DIR / "_data"
SHEETS_DIR = BASE_DIR / "google_sheets"
DIFF_REPORT_PATH = BASE_DIR / "scripts" / "diff_report.xlsx"

# ---------------------------------------------------------------------------
# Files to skip entirely
# ---------------------------------------------------------------------------
SKIP_FILES = {"INDEX.xlsx", "AoR Beginner Game(1).xlsx"}

# ---------------------------------------------------------------------------
# Sheet-name → YAML filename mapping
# ---------------------------------------------------------------------------
SHEET_TO_YAML = {
    "Gear": "gear.yaml",
    "Weapons": "weapons.yaml",
    "Armor": "armor.yaml",
    "Attachments": "attachments.yaml",
    "Qualities": "qualities.yaml",
    "Vehicles": "vehicles.yaml",
    "Starships": "starships.yaml",
    "Vehicles Attachments": "vehicle-attachments.yaml",
    "Vehicle Weapons": "vehicle-weapons.yaml",
    "Additional Rules": "additional-rules.yaml",
    "Skills": "skills.yaml",
    "Talents": "talents.yaml",
    "Abilities": "abilities.yaml",
    "Species": "species.yaml",
    "Adversaries": "adversaries.yaml",
    "Adversaries Gear": "adversaries-gear.yaml",
    "Adversaries Weapons": "adversaries-weapons.yaml",
    "Adversaries Armor": "adversaries-armor.yaml",
    "Creatures": "creatures.yaml",
    "Creatures Weapons": "creatures-weapons.yaml",
    # Variant names
    "Copy of Talents": "talents.yaml",
    "Weapons Attachments": "attachments.yaml",
}

SKIP_SHEETS = {
    "Explosive Charges",
    "Copy from Excel",
    "Sheet25",
    "Summary",
}

# ---------------------------------------------------------------------------
# YAML fields per category (besides index / generatedId / full_name)
# ---------------------------------------------------------------------------
YAML_FIELDS = {
    "weapons.yaml": [
        "category", "skill", "damage", "crit", "range",
        "encumbrance", "hp", "price", "restricted", "rarity", "special",
    ],
    "armor.yaml": [
        "defense", "soak", "price", "restricted", "encumbrance", "hp", "rarity",
    ],
    "gear.yaml": [
        "category", "price", "restricted", "encumbrance", "rarity",
    ],
    "attachments.yaml": [
        "attachment_category", "price", "restricted", "encumbrance", "hp", "rarity",
    ],
    "vehicle-attachments.yaml": [
        "price", "restricted", "hp", "rarity",
    ],
    "vehicle-weapons.yaml": [
        "category", "range", "damage", "crit", "price", "restricted",
        "rarity", "qualities", "compatibleSilhouette", "notes",
    ],
    "vehicles.yaml": [
        "category", "silhouette", "speed", "handling", "model",
        "manufacturer", "crew", "encumbrance", "passengers", "price",
        "restricted", "rarity", "hp", "weapons",
    ],
    "starships.yaml": [
        "category", "silhouette", "speed", "handling", "model",
        "manufacturer", "navicomputer", "crew", "encumbrance", "passengers",
        "price", "restricted", "rarity", "hp", "weapons",
    ],
    "adversaries.yaml": [
        "level", "skills", "talents", "abilities", "equipment",
    ],
    "creatures.yaml": [
        "level", "skills", "talents", "abilities", "equipment",
    ],
    "adversaries-weapons.yaml": [
        "skill", "damage", "crit", "range", "special",
    ],
    "adversaries-armor.yaml": [
        "defense", "soak",
    ],
    "adversaries-gear.yaml": [],
    "creatures-weapons.yaml": [
        "skill", "damage", "crit", "range", "special",
    ],
    "talents.yaml": [
        "activation", "ranked", "forceSensitive",
    ],
    "abilities.yaml": [
        "description",
    ],
    "additional-rules.yaml": [
        "description",
    ],
    "skills.yaml": [
        "characteristic", "type",
    ],
    "species.yaml": [],
    "qualities.yaml": [
        "active", "ranked", "effect",
    ],
}

# Minimum number of data fields (not counting name/index) that must be
# non-empty for a row to be treated as a full item rather than a partial
# book-reference-only row.
MIN_DATA_FIELDS = {
    "adversaries-gear.yaml": 0,
    "species.yaml": 0,
}

# ---------------------------------------------------------------------------
# Column name → YAML field mapping per YAML file
# ---------------------------------------------------------------------------

# Common column mappings shared across many sheets
_COMMON = {
    "name": "full_name",
    "index": "index",
    "generatedid": "generatedId",
}

# Build per-yaml-file column maps
COLUMN_MAP = {}

def _build_column_maps():
    """Construct COLUMN_MAP entries for each YAML file."""

    # weapons
    wm = dict(_COMMON)
    for f in YAML_FIELDS["weapons.yaml"]:
        wm[f] = f
    wm["defence"] = "defense"  # not in weapons but just in case
    COLUMN_MAP["weapons.yaml"] = wm

    # armor
    am = dict(_COMMON)
    for f in YAML_FIELDS["armor.yaml"]:
        am[f] = f
    am["defence"] = "defense"
    COLUMN_MAP["armor.yaml"] = am

    # gear
    gm = dict(_COMMON)
    for f in YAML_FIELDS["gear.yaml"]:
        gm[f] = f
    COLUMN_MAP["gear.yaml"] = gm

    # attachments — spreadsheet "category" maps to "attachment_category"
    atm = dict(_COMMON)
    for f in YAML_FIELDS["attachments.yaml"]:
        atm[f] = f
    atm["category"] = "attachment_category"
    COLUMN_MAP["attachments.yaml"] = atm

    # vehicle-attachments
    vam = dict(_COMMON)
    for f in YAML_FIELDS["vehicle-attachments.yaml"]:
        vam[f] = f
    COLUMN_MAP["vehicle-attachments.yaml"] = vam

    # vehicle-weapons
    vwm = dict(_COMMON)
    for f in YAML_FIELDS["vehicle-weapons.yaml"]:
        vwm[f] = f
    vwm["dam"] = "damage"
    vwm["compatible silhouette"] = "compatibleSilhouette"
    vwm["compatible silhouettes"] = "compatibleSilhouette"
    COLUMN_MAP["vehicle-weapons.yaml"] = vwm

    # vehicles
    vm = dict(_COMMON)
    for f in YAML_FIELDS["vehicles.yaml"]:
        vm[f] = f
    vm["encumbrance capacity"] = "encumbrance"
    COLUMN_MAP["vehicles.yaml"] = vm

    # starships
    sm = dict(_COMMON)
    for f in YAML_FIELDS["starships.yaml"]:
        sm[f] = f
    # "hyperdrive" column is NOT navicomputer — they're separate columns
    # YAML only tracks navicomputer, so "hyperdrive" is intentionally unmapped
    COLUMN_MAP["starships.yaml"] = sm

    # adversaries
    adm = dict(_COMMON)
    for f in YAML_FIELDS["adversaries.yaml"]:
        adm[f] = f
    COLUMN_MAP["adversaries.yaml"] = adm

    # creatures
    cm = dict(_COMMON)
    for f in YAML_FIELDS["creatures.yaml"]:
        cm[f] = f
    COLUMN_MAP["creatures.yaml"] = cm

    # adversaries-weapons
    awm = dict(_COMMON)
    for f in YAML_FIELDS["adversaries-weapons.yaml"]:
        awm[f] = f
    COLUMN_MAP["adversaries-weapons.yaml"] = awm

    # adversaries-armor
    aam = dict(_COMMON)
    for f in YAML_FIELDS["adversaries-armor.yaml"]:
        aam[f] = f
    aam["defence"] = "defense"
    COLUMN_MAP["adversaries-armor.yaml"] = aam

    # adversaries-gear
    agm = dict(_COMMON)
    COLUMN_MAP["adversaries-gear.yaml"] = agm

    # creatures-weapons
    cwm = dict(_COMMON)
    for f in YAML_FIELDS["creatures-weapons.yaml"]:
        cwm[f] = f
    COLUMN_MAP["creatures-weapons.yaml"] = cwm

    # talents
    tm = dict(_COMMON)
    for f in YAML_FIELDS["talents.yaml"]:
        tm[f] = f
    tm["forcesensitive"] = "forceSensitive"
    COLUMN_MAP["talents.yaml"] = tm

    # abilities
    abm = dict(_COMMON)
    for f in YAML_FIELDS["abilities.yaml"]:
        abm[f] = f
    COLUMN_MAP["abilities.yaml"] = abm

    # additional-rules
    arm = dict(_COMMON)
    for f in YAML_FIELDS["additional-rules.yaml"]:
        arm[f] = f
    COLUMN_MAP["additional-rules.yaml"] = arm

    # skills
    skm = dict(_COMMON)
    for f in YAML_FIELDS["skills.yaml"]:
        skm[f] = f
    COLUMN_MAP["skills.yaml"] = skm

    # species
    spm = dict(_COMMON)
    COLUMN_MAP["species.yaml"] = spm

    # qualities
    qm = dict(_COMMON)
    for f in YAML_FIELDS["qualities.yaml"]:
        qm[f] = f
    COLUMN_MAP["qualities.yaml"] = qm


_build_column_maps()

# ---------------------------------------------------------------------------
# UUID regex
# ---------------------------------------------------------------------------
UUID_RE = re.compile(
    r"^[0-9a-f]{8}-[0-9a-f]{4}-[0-9a-f]{4}-[0-9a-f]{4}-[0-9a-f]{12}",
    re.IGNORECASE,
)

# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def normalize_name(name):
    """Lowercase, strip whitespace for matching."""
    if name is None:
        return ""
    return str(name).strip().lower()


def is_uuid_index(index_str):
    """Check if an index string starts with a UUID."""
    return bool(UUID_RE.match(str(index_str).strip()))


def parse_price(value):
    """
    Parse a price value. Handles:
    - Numeric values (int/float)
    - "(R) 2,250" → (2250, True)
    - "2,250" → (2250, False)
    - Regular number → (number, None) — restricted not indicated
    Returns (price_int_or_None, restricted_bool_or_None)
    """
    if value is None:
        return None, None
    if isinstance(value, (int, float)):
        return clean_numeric(value), None
    s = str(value).strip()
    if not s or s == "-":
        return None, None
    restricted = None
    if s.upper().startswith("(R)"):
        restricted = True
        s = s[3:].strip()
    # Remove commas
    s = s.replace(",", "")
    try:
        return int(float(s)), restricted
    except (ValueError, TypeError):
        return None, None


def clean_numeric(value):
    """Convert float-like-int to int, leave others as-is."""
    if value is None:
        return None
    if isinstance(value, float) and value == int(value):
        return int(value)
    return value


def clean_value(value):
    """Clean a cell value for YAML output."""
    if value is None:
        return None
    if isinstance(value, datetime.datetime):
        return None  # Excel date artifacts — not real data
    if isinstance(value, (int, float)):
        return clean_numeric(value)
    s = str(value).strip()
    if not s or s == "-":
        return None
    # Boolean conversions
    if s.lower() in ("yes", "true"):
        return True
    if s.lower() in ("no", "false"):
        return False
    # Try numeric
    try:
        f = float(s.replace(",", ""))
        if f == int(f):
            return int(f)
        return f
    except (ValueError, TypeError):
        pass
    return s


def build_book_name_to_uuid(books_data):
    """Build a map from book full_name (lowercased) → generatedId."""
    mapping = {}
    for book in books_data:
        name = book.get("full_name", "")
        gid = book.get("generatedId", "")
        if name and gid:
            mapping[name.lower()] = gid
    return mapping


def resolve_index_entry(entry_str, book_name_map):
    """
    Resolve an index entry. If it starts with a UUID, return as-is.
    If it starts with a book name, look up the UUID.
    Returns the resolved string or None if unresolvable.
    """
    entry_str = str(entry_str).strip()
    if not entry_str:
        return None

    if is_uuid_index(entry_str):
        return entry_str

    # Format: "Book Name:page"
    parts = entry_str.split(":")
    if len(parts) < 2:
        return None

    page = parts[-1].strip()
    book_name = ":".join(parts[:-1]).strip()
    book_key = book_name.lower()

    if book_key in book_name_map:
        return f"{book_name_map[book_key]}:{page}"

    return None


def get_field_order(yaml_file):
    """Return the field order for new items in a YAML file."""
    fields = list(YAML_FIELDS.get(yaml_file, []))
    fields.append("index")
    fields.append("generatedId")
    fields.append("full_name")
    return fields


def is_partial_row(mapped_fields, yaml_file):
    """
    Determine if a row is a partial book-reference-only row.
    Returns True if the row has too few data fields to be a full item.
    """
    data_fields = YAML_FIELDS.get(yaml_file, [])
    if not data_fields:
        # For categories with no data fields (species, adversaries-gear),
        # any row with a name is considered complete
        return False

    min_required = MIN_DATA_FIELDS.get(yaml_file, 2)
    count = 0
    for f in data_fields:
        if f in mapped_fields and mapped_fields[f] is not None:
            count += 1

    return count < min_required


# ---------------------------------------------------------------------------
# Main processing
# ---------------------------------------------------------------------------

def load_yaml_data(yaml_instance):
    """Load all YAML data files into a dict of {filename: list_of_items}."""
    data = {}
    for yaml_file in YAML_FIELDS:
        path = DATA_DIR / yaml_file
        if path.exists():
            with open(path, "r", encoding="utf-8") as f:
                loaded = yaml_instance.load(f)
                data[yaml_file] = loaded if loaded else []
        else:
            data[yaml_file] = []
    return data


def build_lookup(items):
    """Build a {normalized_name: item} lookup from a list of YAML items."""
    lookup = {}
    for item in items:
        name = normalize_name(item.get("full_name", ""))
        if name:
            lookup[name] = item
    return lookup


def build_index_lookup(items):
    """Build a {normalized_name: set_of_index_entries} for existing items."""
    lookup = {}
    for item in items:
        name = normalize_name(item.get("full_name", ""))
        if name:
            idx = item.get("index", [])
            if idx is None:
                idx = []
            lookup[name] = set(str(e) for e in idx)
    return lookup


def process_spreadsheets(yaml_data, book_name_map, verbose=False):
    """
    Process all XLSX files. Returns:
    - new_items: dict {yaml_file: [list of new item dicts]}
    - diffs: list of (full_name, yaml_file, field, yaml_val, xlsx_val, xlsx_file)
    - skipped: list of (xlsx_file, sheet_name, reason)
    - index_additions: dict {yaml_file: {normalized_name: [new_index_entries]}}
    """
    new_items = {yf: [] for yf in YAML_FIELDS}
    diffs = []
    skipped = []
    index_additions = {yf: {} for yf in YAML_FIELDS}

    # Build lookups per YAML file
    lookups = {yf: build_lookup(items) for yf, items in yaml_data.items()}
    index_lookups = {yf: build_index_lookup(items) for yf, items in yaml_data.items()}

    # Track new items by name to avoid duplicates across spreadsheets
    new_item_names = {yf: set() for yf in YAML_FIELDS}

    xlsx_files = sorted(f for f in os.listdir(SHEETS_DIR) if f.endswith(".xlsx"))

    for xlsx_file in xlsx_files:
        if xlsx_file in SKIP_FILES:
            skipped.append((xlsx_file, "*", "File in skip list"))
            continue

        xlsx_path = SHEETS_DIR / xlsx_file
        if verbose:
            print(f"Processing {xlsx_file}...")

        try:
            wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
        except Exception as e:
            skipped.append((xlsx_file, "*", f"Failed to open: {e}"))
            continue

        for sheet_name in wb.sheetnames:
            if sheet_name in SKIP_SHEETS:
                skipped.append((xlsx_file, sheet_name, "Sheet in skip list"))
                continue

            yaml_file = SHEET_TO_YAML.get(sheet_name)
            if yaml_file is None:
                skipped.append((xlsx_file, sheet_name, "No YAML mapping"))
                continue

            col_map = COLUMN_MAP.get(yaml_file)
            if col_map is None:
                skipped.append((xlsx_file, sheet_name, "No column mapping"))
                continue

            ws = wb[sheet_name]
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                continue

            # Parse headers
            raw_headers = rows[0]
            headers = []
            for h in raw_headers:
                if h is None:
                    headers.append("")
                else:
                    headers.append(str(h).strip().lower().rstrip("\n"))

            # Process data rows
            for row_idx, row in enumerate(rows[1:], start=2):
                if not row:
                    continue

                # Extract name
                name_col_idx = None
                for i, h in enumerate(headers):
                    if h == "name":
                        name_col_idx = i
                        break

                if name_col_idx is None or name_col_idx >= len(row):
                    continue

                raw_name = row[name_col_idx]
                if raw_name is None or str(raw_name).strip() == "":
                    continue

                item_name = str(raw_name).strip()
                norm_name = normalize_name(item_name)

                # Map columns to YAML fields
                mapped = {}
                index_val = None
                has_price_column = False
                restricted_from_price = None

                for i, h in enumerate(headers):
                    if i >= len(row) or not h:
                        continue

                    cell_val = row[i]

                    if h == "index":
                        index_val = cell_val
                        continue

                    # Look up via column map — skip unmapped headers
                    yaml_field = col_map.get(h)
                    if yaml_field is None:
                        continue

                    if yaml_field == "full_name":
                        continue  # Already handled

                    # Special handling for price with (R)
                    if yaml_field == "price":
                        has_price_column = True
                        price_val, r_flag = parse_price(cell_val)
                        if price_val is not None:
                            mapped["price"] = price_val
                        if r_flag is not None:
                            restricted_from_price = r_flag
                        continue

                    if yaml_field == "restricted":
                        val = clean_value(cell_val)
                        if val is not None:
                            mapped["restricted"] = val
                        continue

                    val = clean_value(cell_val)
                    if val is not None:
                        mapped[yaml_field] = val

                # Apply restricted from price if not already set
                if restricted_from_price is not None and "restricted" not in mapped:
                    mapped["restricted"] = restricted_from_price

                # Resolve index
                resolved_indices = []
                if index_val is not None:
                    idx_str = str(index_val).strip()
                    if idx_str:
                        # Some index cells have multiple entries separated by newlines or commas
                        for part in re.split(r"[\n,]+", idx_str):
                            part = part.strip()
                            if part:
                                resolved = resolve_index_entry(part, book_name_map)
                                if resolved:
                                    resolved_indices.append(resolved)

                # Check if this is a partial row
                if is_partial_row(mapped, yaml_file):
                    # Partial row — only use for adding index entries
                    if norm_name in lookups[yaml_file] and resolved_indices:
                        existing_indices = index_lookups[yaml_file].get(norm_name, set())
                        for ri in resolved_indices:
                            if ri not in existing_indices:
                                if norm_name not in index_additions[yaml_file]:
                                    index_additions[yaml_file][norm_name] = []
                                index_additions[yaml_file][norm_name].append(ri)
                                existing_indices.add(ri)
                    if verbose:
                        print(f"  Partial row skipped (index only): {item_name} in {sheet_name}")
                    continue

                # Check against existing YAML data
                existing = lookups[yaml_file].get(norm_name)

                if existing is not None:
                    # Item exists — check for diffs and add index entries
                    allowed_fields = YAML_FIELDS.get(yaml_file, [])
                    for field in allowed_fields:
                        if field not in mapped:
                            continue
                        xlsx_val = mapped[field]
                        yaml_val = existing.get(field)

                        # Normalize for comparison
                        if yaml_val is None and xlsx_val is None:
                            continue

                        # Compare stringified for robustness
                        y_str = str(yaml_val).strip() if yaml_val is not None else ""
                        x_str = str(xlsx_val).strip() if xlsx_val is not None else ""

                        if y_str.lower() != x_str.lower():
                            diffs.append((
                                item_name, yaml_file, field,
                                yaml_val, xlsx_val, xlsx_file
                            ))

                    # Add new index entries
                    if resolved_indices:
                        existing_indices = index_lookups[yaml_file].get(norm_name, set())
                        for ri in resolved_indices:
                            if ri not in existing_indices:
                                if norm_name not in index_additions[yaml_file]:
                                    index_additions[yaml_file][norm_name] = []
                                index_additions[yaml_file][norm_name].append(ri)
                                existing_indices.add(ri)
                else:
                    # New item
                    if norm_name in new_item_names[yaml_file]:
                        # Already added from another spreadsheet — just add index
                        for ni in new_items[yaml_file]:
                            if normalize_name(ni.get("full_name", "")) == norm_name:
                                if resolved_indices:
                                    existing_idx = set(str(e) for e in ni.get("index", []))
                                    for ri in resolved_indices:
                                        if ri not in existing_idx:
                                            ni["index"].append(ri)
                                            existing_idx.add(ri)
                                break
                        continue

                    new_item = {}
                    field_order = get_field_order(yaml_file)
                    for field in field_order:
                        if field == "full_name":
                            new_item["full_name"] = item_name
                        elif field == "generatedId":
                            new_item["generatedId"] = str(uuid.uuid4())
                        elif field == "index":
                            new_item["index"] = resolved_indices if resolved_indices else []
                        elif field in mapped:
                            new_item[field] = mapped[field]

                    new_items[yaml_file].append(new_item)
                    new_item_names[yaml_file].add(norm_name)

                    if verbose:
                        print(f"  New item: {item_name} → {yaml_file}")

        wb.close()

    return new_items, diffs, skipped, index_additions


def write_yaml_files(yaml_instance, yaml_data, new_items, index_additions, dry_run=False, verbose=False):
    """Write updated YAML files with new items appended and new index entries added."""
    for yaml_file, items in yaml_data.items():
        new = new_items.get(yaml_file, [])
        idx_adds = index_additions.get(yaml_file, {})

        if not new and not idx_adds:
            continue

        # Add new index entries to existing items
        if idx_adds:
            for item in items:
                norm = normalize_name(item.get("full_name", ""))
                if norm in idx_adds:
                    idx = item.get("index")
                    if idx is None:
                        item["index"] = []
                        idx = item["index"]
                    existing_set = set(str(e) for e in idx)
                    for new_idx in idx_adds[norm]:
                        if new_idx not in existing_set:
                            idx.append(new_idx)
                            existing_set.add(new_idx)
                    if verbose:
                        print(f"  Added {len(idx_adds[norm])} index entries to {item.get('full_name')} in {yaml_file}")

        # Append new items
        if new:
            items.extend(new)
            if verbose:
                print(f"  Appended {len(new)} new items to {yaml_file}")

        if dry_run:
            print(f"[DRY RUN] Would write {yaml_file}: {len(new)} new items, {len(idx_adds)} index updates")
            continue

        path = DATA_DIR / yaml_file
        with open(path, "w", encoding="utf-8") as f:
            yaml_instance.dump(items, f)

        if verbose:
            print(f"  Wrote {yaml_file}")


def write_diff_report(diffs, new_items, skipped):
    """Write the diff report XLSX."""
    wb = openpyxl.Workbook()

    # Diffs sheet
    ws_diffs = wb.active
    ws_diffs.title = "Diffs"
    ws_diffs.append(["full_name", "category", "field", "yaml_value", "xlsx_value", "xlsx_file"])
    for full_name, yaml_file, field, yaml_val, xlsx_val, xlsx_file in diffs:
        category = yaml_file.replace(".yaml", "")
        ws_diffs.append([full_name, category, field, str(yaml_val) if yaml_val is not None else "", str(xlsx_val) if xlsx_val is not None else "", xlsx_file])

    # New Items sheet
    ws_new = wb.create_sheet("New Items")
    ws_new.append(["category", "full_name", "generatedId", "index", "xlsx_source"])
    for yaml_file, items in new_items.items():
        category = yaml_file.replace(".yaml", "")
        for item in items:
            idx_str = ", ".join(str(e) for e in item.get("index", []))
            ws_new.append([category, item.get("full_name", ""), item.get("generatedId", ""), idx_str, ""])

    # Skipped sheet
    ws_skip = wb.create_sheet("Skipped")
    ws_skip.append(["xlsx_file", "sheet_name", "reason"])
    for xlsx_file, sheet_name, reason in skipped:
        ws_skip.append([xlsx_file, sheet_name, reason])

    wb.save(DIFF_REPORT_PATH)
    print(f"Diff report written to {DIFF_REPORT_PATH}")


def main():
    parser = argparse.ArgumentParser(description="Merge XLSX data into YAML files")
    parser.add_argument("--dry-run", action="store_true",
                        help="Process everything but don't modify YAML files")
    parser.add_argument("--verbose", action="store_true",
                        help="Detailed logging")
    args = parser.parse_args()

    yaml_instance = YAML()
    yaml_instance.preserve_quotes = True
    yaml_instance.default_flow_style = False
    yaml_instance.width = 4096  # Avoid line wrapping
    yaml_instance.explicit_start = True  # Preserve --- document start

    # Load books for name→UUID resolution
    books_path = DATA_DIR / "books.yaml"
    with open(books_path, "r", encoding="utf-8") as f:
        books_data = yaml_instance.load(f)
    book_name_map = build_book_name_to_uuid(books_data)

    if args.verbose:
        print(f"Loaded {len(book_name_map)} book name→UUID mappings")

    # Load all YAML data
    yaml_data = load_yaml_data(yaml_instance)

    if args.verbose:
        for yf, items in yaml_data.items():
            print(f"  {yf}: {len(items)} existing items")

    # Process spreadsheets
    new_items, diffs, skipped, index_additions = process_spreadsheets(
        yaml_data, book_name_map, verbose=args.verbose
    )

    # Summary
    total_new = sum(len(v) for v in new_items.values())
    total_idx = sum(len(v) for v in index_additions.values())
    print(f"\nSummary:")
    print(f"  New items found: {total_new}")
    print(f"  Items with new index entries: {total_idx}")
    print(f"  Field diffs on existing items: {len(diffs)}")
    print(f"  Skipped sheets: {len(skipped)}")

    for yf in sorted(YAML_FIELDS.keys()):
        n = len(new_items.get(yf, []))
        idx = len(index_additions.get(yf, {}))
        if n or idx:
            print(f"    {yf}: {n} new, {idx} index updates")

    # Write diff report
    write_diff_report(diffs, new_items, skipped)

    # Write YAML files
    write_yaml_files(yaml_instance, yaml_data, new_items, index_additions,
                     dry_run=args.dry_run, verbose=args.verbose)

    if args.dry_run:
        print("\n[DRY RUN] No YAML files were modified.")


if __name__ == "__main__":
    main()
